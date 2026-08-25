#!/usr/bin/env python3
"""
import_xlsx.py — Convierte el libro SDG_Arrival_Charges (.xlsx exportado de
Google Sheets) en los CSV canónicos de data/.

Uso:
    python scripts/import_xlsx.py ruta/al/SDG_Arrival_Charges.xlsx

Solo se usa cuando alguien edita el Google Sheet y quiere volcarlo al repo.
El día a día (añadir un aeropuerto) se hace editando data/*.csv directamente.
"""
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"

AIRPORT_COLS = ["code", "name", "country", "region", "active"]

CHARGE_COLS = [
    "code",
    "currency",
    "sec_min", "sec_rate",
    "cus_min", "cus_rate",
    "truck_bulk_min", "truck_bulk_fee_mawb", "truck_bulk_rate",
    "truck_uld_min", "truck_uld_fee",
    "thc_gen_min", "thc_gen_rate",
    "thc_dg_min", "thc_dg_rate",
    "thc_pha_min", "thc_pha_rate",
    "st_gen_mawb_fee", "st_gen_rate_1_20", "st_gen_rate_21plus", "st_gen_free_days",
    "st_cool_mawb_fee", "st_cool_rate_1_20", "st_cool_rate_21plus", "st_cool_free_days",
    "st_dg_mawb_fee", "st_dg_rate_1_20", "st_dg_rate_21plus", "st_dg_free_days",
    "imp_doc_min", "imp_doc_fee",
]

# Índices de columna (0-based) en la hoja "Arrival Charges Table"
CHARGE_MAP = {
    "sec_min": 1, "sec_rate": 2,
    "cus_min": 4, "cus_rate": 5,
    "truck_bulk_min": 7, "truck_bulk_fee_mawb": 8, "truck_bulk_rate": 9,
    "truck_uld_min": 11, "truck_uld_fee": 12,
    "thc_gen_min": 14, "thc_gen_rate": 15,
    "thc_dg_min": 17, "thc_dg_rate": 18,
    "thc_pha_min": 20, "thc_pha_rate": 21,
    "st_gen_mawb_fee": 23, "st_gen_rate_1_20": 24, "st_gen_rate_21plus": 25, "st_gen_free_days": 28,
    "st_cool_mawb_fee": 29, "st_cool_rate_1_20": 30, "st_cool_rate_21plus": 31, "st_cool_free_days": 34,
    "st_dg_mawb_fee": 35, "st_dg_rate_1_20": 36, "st_dg_rate_21plus": 37, "st_dg_free_days": 40,
    "imp_doc_min": 41, "imp_doc_fee": 42,
}


def num(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if s in ("", "-", "N/A", "n/a"):
        return default
    try:
        return float(s)
    except ValueError:
        return default


def main(path):
    wb = load_workbook(path, read_only=True, data_only=True)

    # ── Airports ──────────────────────────────────────────────
    rows = list(wb["Airports"].iter_rows(values_only=True))
    airports = []
    for r in rows[5:]:
        if not r or not r[0]:
            continue
        airports.append({
            "code": str(r[0]).strip(),
            "name": str(r[1]).strip(),
            "country": str(r[3]).strip() if len(r) > 3 and r[3] else "",
            "region": str(r[4]).strip() if len(r) > 4 and r[4] else "",
            "active": "yes" if str(r[5]).strip().lower() in ("yes", "sí", "si", "true", "1") else "no",
        })

    with (DATA / "airports.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=AIRPORT_COLS, lineterminator="\n")
        w.writeheader()
        w.writerows(sorted(airports, key=lambda a: a["code"]))

    # ── Arrival charges ───────────────────────────────────────
    rows = list(wb["Arrival Charges Table"].iter_rows(values_only=True))
    charges = []
    for r in rows[5:]:
        if not r or not r[0]:
            continue
        label = str(r[0])
        code = label.split("-")[0].strip()
        rec = {"code": code, "currency": "EUR"}
        for key, idx in CHARGE_MAP.items():
            rec[key] = round(num(r[idx] if len(r) > idx else None), 4)
        charges.append(rec)

    with (DATA / "arrival_charges.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CHARGE_COLS, lineterminator="\n")
        w.writeheader()
        w.writerows(sorted(charges, key=lambda c: c["code"]))

    print(f"OK  {len(airports)} aeropuertos → data/airports.csv")
    print(f"OK  {len(charges)} tarifas de llegada → data/arrival_charges.csv")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Uso: python scripts/import_xlsx.py <fichero.xlsx>")
    main(sys.argv[1])
